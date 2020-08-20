import openpyxl as opx

def getRowCount(file, client):
    sheet = client.open(file).sheet1
    return len(sheet.get_all_values())

def getColumnCount(file, client):
    sheet = client.open(file).sheet1
    return len(sheet.get_all_values()[0])

def readData(file, client, rowno, columno):
    sheet = client.open(file).sheet1
    return sheet.cell(row=rowno, column=columno).value

def writeData(file, client, rowno, columno, data):
    sheet = client.open(file).sheet1
    sheet.update_cell(rowno, columno, data)
    # workbook = opx.load_workbook(file)
    # sheet = workbook.worksheets[sheet]  # 0 -int
    # sheet.cell(row=rowno, column=columno).value = data
    # workbook.save(file)