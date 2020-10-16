import openpyxl as opx

def getRowCount(file, sheet):
    workbook = opx.load_workbook(file)
    sheet = workbook.worksheets[sheet] #0 -int
    return sheet.max_row

def getColumnCount(file, sheet):
    workbook = opx.load_workbook(file)
    sheet = workbook.worksheets[sheet]  # 0 -int
    return sheet.max_column

def readData(file, sheet, rowno, columno):
    workbook = opx.load_workbook(file)
    sheet = workbook.worksheets[sheet]  # 0 -int
    return sheet.cell(row=rowno, column=columno).value

def writeData(file, sheet, rowno, columno, data):
    workbook = opx.load_workbook(file)
    sheet = workbook.worksheets[sheet]  # 0 -int
    sheet.cell(row=rowno, column=columno).value = data
    workbook.save(file)